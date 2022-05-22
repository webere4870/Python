import openpyxl



workbook = openpyxl.load_workbook("tester.xlsx")
formattedWorkbook = openpyxl.load_workbook("formatted.xlsx")
sheet = workbook.active
max = sheet.max_row
newSheet = formattedWorkbook.active
arrayOfObjects = []
dictionary = {}



for i in range(1, max + 1):
    val=sheet.cell(row=i, column=1)
    obj = val.value
    stripper = obj.replace('"', '')
    individuals = stripper.split(",")
    newObj = {}
    for temp in individuals:
        finalSplit = temp.split(':')
        key = finalSplit[0]
        val = finalSplit[1]
        newObj[key] = val
        if dictionary.get(key):
            dictionary[key]=dictionary[key]+1
        else:
            dictionary[key]=1
    arrayOfObjects.append(newObj)



loopCount = 1
for key, val in dictionary.items():
    cell=newSheet.cell(row=1, column=loopCount)
    cell.value = key
    loopCount += 1

print(arrayOfObjects)

workbook.save("tester.xlsx")
formattedWorkbook.save("formatted.xlsx")
#formattedWorkbook.save()
